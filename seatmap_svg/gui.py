from __future__ import annotations

import argparse
import os
import threading
import traceback
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from .config import load_config
from .pipeline import run_pipeline

DEFAULT_CONFIG_PATH = "configs/default.yaml"


class SeatmapSVGApp(tk.Tk):
    """Tkinter GUI wrapper around the shared Excel-to-SVG pipeline."""

    def __init__(self, config_path: str = DEFAULT_CONFIG_PATH) -> None:
        super().__init__()
        self.title("Seatmap SVG Tool")
        self.geometry("860x620")
        self.minsize(760, 560)
        self.last_output_dir = Path("output")

        config = self._safe_load_config(config_path)
        self.config_path_var = tk.StringVar(value=config_path)
        self.excel_path_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value=config.get("input", {}).get("sheet", ""))
        self.range_var = tk.StringVar(value=config.get("input", {}).get("range", "A1:Z50"))
        self.svg_path_var = tk.StringVar(value=config.get("output", {}).get("svg", "output/seatmap.svg"))
        self.report_path_var = tk.StringVar(value=config.get("output", {}).get("report", "output/report.json"))
        self.status_var = tk.StringVar(value="请选择 Excel 文件后开始生成。")

        self._build_widgets()

    def _safe_load_config(self, config_path: str) -> dict:
        try:
            return load_config(config_path)
        except Exception:
            return load_config(None)

    def _build_widgets(self) -> None:
        root = ttk.Frame(self, padding=12)
        root.pack(fill=tk.BOTH, expand=True)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(8, weight=1)

        ttk.Label(root, text="配置文件").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(root, textvariable=self.config_path_var).grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="选择配置文件", command=self.choose_config).grid(row=0, column=2, sticky="ew")

        ttk.Label(root, text="Excel 文件").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(root, textvariable=self.excel_path_var).grid(row=1, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="选择 Excel 文件", command=self.choose_excel).grid(row=1, column=2, sticky="ew")

        ttk.Label(root, text="Sheet").grid(row=2, column=0, sticky="w", pady=4)
        self.sheet_combo = ttk.Combobox(root, textvariable=self.sheet_var, state="readonly")
        self.sheet_combo.grid(row=2, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="刷新 Sheet", command=self.load_sheets).grid(row=2, column=2, sticky="ew")

        ttk.Label(root, text="扫描范围").grid(row=3, column=0, sticky="w", pady=4)
        ttk.Entry(root, textvariable=self.range_var).grid(row=3, column=1, sticky="ew", padx=8)
        ttk.Label(root, text="例如 A1:Z50").grid(row=3, column=2, sticky="w")

        ttk.Label(root, text="SVG 输出").grid(row=4, column=0, sticky="w", pady=4)
        ttk.Entry(root, textvariable=self.svg_path_var).grid(row=4, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="选择 SVG 输出位置", command=self.choose_svg).grid(row=4, column=2, sticky="ew")

        ttk.Label(root, text="JSON 报告").grid(row=5, column=0, sticky="w", pady=4)
        ttk.Entry(root, textvariable=self.report_path_var).grid(row=5, column=1, sticky="ew", padx=8)
        ttk.Button(root, text="选择报告输出位置", command=self.choose_report).grid(row=5, column=2, sticky="ew")

        actions = ttk.Frame(root)
        actions.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(10, 6))
        self.generate_button = ttk.Button(actions, text="开始生成", command=self.start_generate)
        self.generate_button.pack(side=tk.LEFT)
        ttk.Button(actions, text="打开输出目录", command=self.open_output_dir).pack(side=tk.LEFT, padx=8)
        ttk.Button(actions, text="打开 SVG 文件", command=self.open_svg).pack(side=tk.LEFT)

        ttk.Label(root, textvariable=self.status_var).grid(row=7, column=0, columnspan=3, sticky="w", pady=4)

        log_frame = ttk.LabelFrame(root, text="日志")
        log_frame.grid(row=8, column=0, columnspan=3, sticky="nsew")
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)
        self.log_text = tk.Text(log_frame, height=14, wrap="word")
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

    def log(self, message: str) -> None:
        def append() -> None:
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.status_var.set(message)
        self.after(0, append)

    def choose_config(self) -> None:
        path = filedialog.askopenfilename(title="选择配置文件", filetypes=[("YAML", "*.yaml *.yml"), ("所有文件", "*.*")])
        if path:
            self.config_path_var.set(path)
            try:
                config = load_config(path)
                self.range_var.set(config.get("input", {}).get("range", self.range_var.get()))
                self.svg_path_var.set(config.get("output", {}).get("svg", self.svg_path_var.get()))
                self.report_path_var.set(config.get("output", {}).get("report", self.report_path_var.get()))
                self.log(f"已加载配置文件: {path}")
            except Exception as exc:
                self.log(f"配置文件读取失败: {exc}")
                messagebox.showerror("配置文件读取失败", str(exc))

    def choose_excel(self) -> None:
        path = filedialog.askopenfilename(title="选择 Excel 文件", filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")])
        if path:
            self.excel_path_var.set(path)
            self.load_sheets()

    def choose_svg(self) -> None:
        path = filedialog.asksaveasfilename(title="选择 SVG 输出位置", defaultextension=".svg", filetypes=[("SVG", "*.svg")], initialfile=Path(self.svg_path_var.get()).name)
        if path:
            self.svg_path_var.set(path)

    def choose_report(self) -> None:
        path = filedialog.asksaveasfilename(title="选择 JSON 报告输出位置", defaultextension=".json", filetypes=[("JSON", "*.json")], initialfile=Path(self.report_path_var.get()).name)
        if path:
            self.report_path_var.set(path)

    def load_sheets(self) -> None:
        excel_path = self.excel_path_var.get().strip()
        if not excel_path:
            messagebox.showwarning("未选择 Excel 文件", "请先选择 Excel 文件。")
            return
        if not Path(excel_path).exists():
            messagebox.showerror("Excel 文件不存在", excel_path)
            return
        try:
            self.log("正在读取 Sheet 列表")
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            self.sheet_combo["values"] = sheets
            if sheets:
                self.sheet_var.set(sheets[0])
            self.log(f"已读取 {len(sheets)} 个 Sheet。")
        except Exception as exc:
            self.log(f"读取 Sheet 失败: {exc}")
            messagebox.showerror("读取 Sheet 失败", str(exc))

    def collect_params(self) -> dict:
        return {
            "config_path": self.config_path_var.get().strip() or DEFAULT_CONFIG_PATH,
            "input_path": self.excel_path_var.get().strip(),
            "sheet": self.sheet_var.get().strip(),
            "scan_range": self.range_var.get().strip(),
            "output_path": self.svg_path_var.get().strip(),
            "report_path": self.report_path_var.get().strip(),
        }

    def start_generate(self) -> None:
        params = self.collect_params()
        if not params["input_path"]:
            messagebox.showwarning("未选择 Excel 文件", "请先选择 Excel 文件。")
            return
        self.generate_button.configure(state=tk.DISABLED)
        self.log_text.delete("1.0", tk.END)
        thread = threading.Thread(target=self._generate_worker, args=(params,), daemon=True)
        thread.start()

    def _generate_worker(self, params: dict) -> None:
        try:
            result = run_pipeline(log_callback=self.log, **params)
            self.last_output_dir = Path(result["svg"]).parent
            self.after(0, lambda: messagebox.showinfo("生成成功", f"SVG: {result['svg']}\n报告: {result['report']}"))
        except Exception as exc:
            detail = traceback.format_exc()
            self.log(f"生成失败: {exc}")
            self.log(detail)
            self.after(0, lambda: messagebox.showerror("生成失败", str(exc)))
        finally:
            self.after(0, lambda: self.generate_button.configure(state=tk.NORMAL))

    def open_output_dir(self) -> None:
        path = Path(self.svg_path_var.get() or self.last_output_dir).parent
        path.mkdir(parents=True, exist_ok=True)
        self._open_path(path)

    def open_svg(self) -> None:
        path = Path(self.svg_path_var.get())
        if not path.exists():
            messagebox.showwarning("SVG 不存在", f"文件不存在: {path}")
            return
        self._open_path(path)

    def _open_path(self, path: Path) -> None:
        try:
            os.startfile(str(path))  # type: ignore[attr-defined]
        except AttributeError:
            import subprocess
            subprocess.Popen(["open" if os.name == "posix" else "xdg-open", str(path)])
        except Exception as exc:
            messagebox.showerror("打开失败", str(exc))


def main(argv=None) -> None:
    parser = argparse.ArgumentParser(description="Launch Seatmap SVG Tool GUI")
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="Initial YAML config path")
    args = parser.parse_args(argv)
    app = SeatmapSVGApp(config_path=args.config)
    app.mainloop()
