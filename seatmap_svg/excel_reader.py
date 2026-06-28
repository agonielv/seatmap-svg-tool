from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from .xlcoords import get_column_letter, range_boundaries

@dataclass
class CellInfo:
    row: int
    col: int
    coordinate: str
    value: str
    fill_key: str
    fill_rgb: Optional[str]
    is_merged: bool
    merge_range: Optional[str]
    merge_bounds: Optional[Tuple[int, int, int, int]]

@dataclass
class SheetData:
    file: str
    sheet: str
    scan_range: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int
    cells: List[CellInfo]
    merged_ranges: List[str]

def fill_to_key(cell) -> tuple[str, Optional[str]]:
    fill = cell.fill
    fg = fill.fgColor
    if not fill or fill.fill_type is None:
        return "none", None
    if fg.type == "rgb" and fg.rgb:
        rgb = fg.rgb[-6:].upper()
        return f"rgb:{rgb}", f"#{rgb}"
    if fg.type == "theme":
        return f"theme:{fg.theme}", None
    if fg.type == "indexed":
        return f"indexed:{fg.indexed}", None
    return f"{fg.type}:{fg.value}", None

def _merge_lookup(ws) -> Dict[str, Tuple[str, Tuple[int, int, int, int]]]:
    lookup: Dict[str, Tuple[str, Tuple[int, int, int, int]]] = {}
    for merged in ws.merged_cells.ranges:
        bounds = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                lookup[f"{get_column_letter(col)}{row}"] = (str(merged), bounds)
    return lookup

def read_sheet(file: str, sheet: str | None, scan_range: str) -> SheetData:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required to read Excel files. Install with: python -m pip install -r requirements.txt") from exc
    wb = load_workbook(filename=file, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    min_col, min_row, max_col, max_row = range_boundaries(scan_range)
    merge_lookup = _merge_lookup(ws)
    cells: List[CellInfo] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            key, rgb = fill_to_key(cell)
            coord = cell.coordinate
            merge = merge_lookup.get(coord)
            cells.append(CellInfo(row, col, coord, "" if cell.value is None else str(cell.value).strip(), key, rgb, bool(merge), merge[0] if merge else None, merge[1] if merge else None))
    return SheetData(str(Path(file)), ws.title, scan_range, min_col, min_row, max_col, max_row, cells, [str(r) for r in ws.merged_cells.ranges])
